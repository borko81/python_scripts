from openpyxl import load_workbook
FILE = 'work.xlsx'

workbook = load_workbook(FILE)
sheet = workbook.active

print(sheet.title)

for line in sheet["A1:B3"]:
    for cell in line:
        print(cell.value)

print("-" * 20)

for line in sheet["A"]:
    print(line.value)

print("-" * 20)

for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)